import json
import logging
from datetime import datetime, timedelta
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group, User
from django.core.paginator import Paginator
from django.db.models import Max, Min, Sum, Q
from django.utils.http import urlencode
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render 
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font

from .forms import ProductoForm
from .models import (Categoria, Comanda, DetalleComanda, EliminacionComanda,
                    HistorialComanda, Producto, Cliente, DireccionCliente)
from core.utils.payments import PaymentError, calculate_payment_breakdown
from core.utils.printers import PrinterError, send_raw_to_printer


logger = logging.getLogger(__name__)

OWNER_GROUP_NAME = 'Dueños'
VALID_ESTADOS_COMANDA = {'abierta', 'cerrada'}
VALID_TIPOS_SERVICIO = {'servir', 'delivery', 'reserva'}


def _user_is_owner(user) -> bool:
    return user.is_superuser or user.groups.filter(name=OWNER_GROUP_NAME).exists()


def _sanitize_estado(value: str | None) -> str:
    estado = (value or 'abierta').lower()
    return estado if estado in VALID_ESTADOS_COMANDA else 'abierta'


def _sanitize_tipo_servicio(value: str | None) -> str:
    tipo = (value or 'servir').lower()
    return tipo if tipo in VALID_TIPOS_SERVICIO else 'servir'


def _next_numero_cliente_servir() -> int:
    max_numero = Comanda.objects.filter(
        tipo_servicio='servir',
        numero_cliente__isnull=False,
    ).aggregate(max_numero=Max('numero_cliente'))['max_numero']
    # Si no hay registros previos, comenzar en 1 en lugar de 0
    return 1 if max_numero is None else max_numero + 1


def _resolve_cliente_data(
    data: dict,
    tipo_servicio: str,
    comanda: Comanda | None = None,
) -> tuple[str, str | None, str | None, int | None]:
    """Devuelve (nombre_cliente, telefono_cliente, direccion_cliente, numero_cliente)."""
    nombre = (data.get('cliente') or '').strip()
    telefono = (data.get('telefono_cliente') or '').strip()
    direccion = (data.get('direccion_cliente') or '').strip()

    if tipo_servicio == 'servir':
        # Mantener numero si ya existe al editar; si no, asignar siguiente secuencial.
        numero = comanda.numero_cliente if comanda and comanda.numero_cliente is not None else _next_numero_cliente_servir()
        return f'Cliente {numero}', None, None, numero

    if not nombre:
        raise ValueError('Debe ingresar el nombre del cliente para delivery o reserva')

    if not telefono:
        raise ValueError('Debe ingresar el teléfono para delivery o reserva')

    if not direccion:
        raise ValueError('Debe ingresar la dirección para delivery o reserva')

    return nombre, telefono, direccion, None


@login_required(login_url='login')
def buscar_cliente_por_telefono(request):
    telefono = (request.GET.get('telefono') or '').strip()
    if not telefono:
        return JsonResponse({'encontrado': False})

    cliente = (
        Cliente.objects
        .prefetch_related('direcciones')
        .filter(telefono=telefono)
        .order_by('-id')
        .first()
    )

    if not cliente:
        return JsonResponse({'encontrado': False})

    direccion = cliente.direcciones.order_by('id').values_list('direccion', flat=True).first() or ''
    direcciones = list(cliente.direcciones.order_by('id').values_list('direccion', flat=True))
    return JsonResponse({
        'encontrado': True,
        'cliente': {
            'id': cliente.id,
            'nombre': cliente.nombre,
            'telefono': cliente.telefono,
            'direccion': direccion,
            'direcciones': direcciones,
        }
    })


def _build_detalle_payload(productos_payload):
    if not productos_payload:
        raise ValueError('Debe agregar al menos un producto a la comanda')

    try:
        producto_ids = {int(item['productoId']) for item in productos_payload}
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError('Formato de productos inválido') from exc

    productos_map = Producto.objects.in_bulk(producto_ids)
    detalles = []
    total = 0

    for item in productos_payload:
        producto_id = int(item['productoId'])
        producto = productos_map.get(producto_id)
        if producto is None:
            raise ValueError(f'Producto con ID {producto_id} no existe')

        try:
            cantidad = max(int(item.get('cantidad', 1)), 1)
        except (TypeError, ValueError) as exc:
            raise ValueError('La cantidad debe ser un entero positivo') from exc

        subtotal = cantidad * producto.precio
        total += subtotal
        nota_producto = (item.get('nota') or '').strip()
        detalles.append({
            'producto': producto,
            'cantidad': cantidad,
            'subtotal': subtotal,
            'nota': nota_producto or None,
        })

    return detalles, total

# Autenticación
def login_vista(request):
    if request.method == 'POST':
        usuario = request.POST.get('username')
        contra = request.POST.get('password')
        user = authenticate(request, username=usuario, password=contra)
        if user is not None:
            auth_login(request, user)
            return redirect('inicio')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'core/login.html')

def registro(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            empleados_group, created = Group.objects.get_or_create(name="Empleados")
            user.groups.add(empleados_group)
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'core/registro.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login') 

# Vistas principales
@login_required(login_url='login')
def inicio(request):
    usuario = request.user
    hora_actual = timezone.now()
    hoy = hora_actual.date()

    comandas_hoy = Comanda.objects.filter(estado='cerrada', fecha__date=hoy)
    detalles = DetalleComanda.objects.filter(comanda__in=comandas_hoy)

    dict_ventas = defaultdict(int)
    # Ventas activas por producto (usamos producto__id y producto__nombre)
    for v in detalles.values('producto__id', 'producto__nombre').annotate(total_vendido=Sum('cantidad')):
        dict_ventas[(v['producto__id'], v['producto__nombre'])] += v['total_vendido']

    historiales_hoy = HistorialComanda.objects.filter(fecha__date=hoy)
    for h in historiales_hoy:
        try:
            detalles_historial = json.loads(h.detalles)
            for item in detalles_historial:
                nombre_producto = item.get('producto')
                cantidad = item.get('cantidad', 0)
                if nombre_producto and cantidad:
                    # Usamos solo el nombre como clave porque no tenemos id
                    # Para evitar mezclar con productos que tienen el mismo nombre, puedes cambiar
                    # a solo usar nombre o implementar lógica para buscar id si quieres
                    dict_ventas[(None, nombre_producto)] += cantidad
        except Exception:
            pass

    # Combinar ventas por nombre (sumar ambos)
    resumen_por_nombre = defaultdict(int)
    for key, cant in dict_ventas.items():
        nombre = key[1]
        resumen_por_nombre[nombre] += cant

    # Ordenar top 3 por cantidad
    top_productos = sorted(
        [{'producto__nombre': nombre, 'total_vendido': cant} for nombre, cant in resumen_por_nombre.items()],
        key=lambda x: x['total_vendido'],
        reverse=True
    )[:3]

    return render(request, 'core/home.html', {
        'usuario': usuario,
        'hora_actual': hora_actual,
        'top_productos': top_productos,
    })

#Usuarios
@login_required(login_url='login')
def usuarios(request):
    if not _user_is_owner(request.user):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('inicio')
    usuarios_qs = User.objects.all().order_by('date_joined')
    paginator = Paginator(usuarios_qs, 10)
    page_number = request.GET.get('page')
    usuarios = paginator.get_page(page_number)
    total_usuarios = paginator.count
    return render(request, 'core/usuarios.html', {
        'usuarios': usuarios,
        'total_usuarios': total_usuarios
    })


@login_required(login_url='login')
def comandas_por_usuario(request, user_id):
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        usuario = User.objects.get(pk=user_id)
        comandas = Comanda.objects.filter(empleado_id=user_id).order_by('-fecha')
        hoy = timezone.now().date()
        comandas_hoy = comandas.filter(fecha__date=hoy)
        total_vendido_hoy = sum(c.total for c in comandas_hoy)
        cantidad_hoy = comandas_hoy.count()

        data = [{
            'id': c.id,
            'nombre_cliente': c.nombre_cliente,
            'fecha': c.fecha.strftime('%d/%m/%Y %H:%M'),
            'estado': c.estado,
            'total': c.total
        } for c in comandas_hoy]

        return JsonResponse({
            'comandas': data,
            'cantidad_hoy': cantidad_hoy,
            'total_vendido_hoy': total_vendido_hoy,
            'last_login': usuario.last_login.strftime('%d/%m/%Y %H:%M') if usuario.last_login else 'Nunca',
            'is_active': usuario.is_active
        })
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
@csrf_exempt
@login_required(login_url='login')
def cambiar_estado_usuario(request, user_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    activar = data.get('activar')
    if activar is None:
        return JsonResponse({'error': 'Falta el parámetro "activar"'}, status=400)

    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    user.is_active = bool(activar)
    user.save(update_fields=['is_active'])
    return JsonResponse({'success': True})


@csrf_exempt
@login_required(login_url='login')
def editar_usuario(request, user_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    username = (data.get('username') or '').strip()
    password = data.get('password')
    grupo_nombre = data.get('grupo')

    if not username or not password or not grupo_nombre:
        return JsonResponse({'success': False, 'error': 'Faltan datos obligatorios'})

    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    user.username = username
    user.password = make_password(password)
    user.save()

    grupo = Group.objects.filter(name=grupo_nombre).first()
    if not grupo:
        return JsonResponse({'success': False, 'error': f'Grupo "{grupo_nombre}" no encontrado'})

    user.groups.clear()
    user.groups.add(grupo)
    return JsonResponse({'success': True})


@csrf_exempt
@login_required(login_url='login')
def eliminar_usuario(request, user_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    if user == request.user:
        return JsonResponse({'error': 'No puedes eliminar tu propio usuario'}, status=400)

    user.delete()
    return JsonResponse({'success': True})



@login_required(login_url='login')
def venta(request):
    productos = Producto.objects.all()
    comandas = Comanda.objects.all().order_by('-fecha')
    categorias = Categoria.objects.all()
    return render(request, 'core/venta.html', {
        'productos': productos,
        'comandas': comandas,
        'categorias': categorias,
        'es_owner_actual': _user_is_owner(request.user),
    })


@login_required(login_url='login')
def reportes(request):
    if not _user_is_owner(request.user):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('inicio')

    filtro = request.GET.get('filtro', 'diario')
    busqueda = (request.GET.get('q') or '').strip()
    metodo_actual = (request.GET.get('metodo_pago') or '').strip()
    orden = request.GET.get('orden') or 'fecha_desc'
    hoy = timezone.localdate()
    historial_queryset = None

    if filtro == 'diario':
        historial_queryset = HistorialComanda.objects.filter(fecha__date=hoy)
    elif filtro == 'semanal':
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        historial_queryset = HistorialComanda.objects.filter(fecha__date__gte=inicio_semana)
    elif filtro == 'mensual':
        historial_queryset = HistorialComanda.objects.filter(
            fecha__year=hoy.year,
            fecha__month=hoy.month
        )
    elif filtro == 'anual':
        historial_queryset = HistorialComanda.objects.filter(fecha__year=hoy.year)
    else:
        historial_queryset = HistorialComanda.objects.none()

    if busqueda:
        historial_queryset = historial_queryset.filter(
            Q(nombre_cliente__icontains=busqueda) |
            Q(empleado__username__icontains=busqueda)
        )

    if metodo_actual:
        historial_queryset = historial_queryset.filter(metodo_pago=metodo_actual)

    # Total de ventas del queryset filtrado (sin paginación)
    total_ventas = historial_queryset.aggregate(total=Sum('total'))['total'] or 0

    ordering_map = {
        'fecha_asc': 'fecha',
        'fecha_desc': '-fecha',
        'total_asc': 'total',
        'total_desc': '-total',
    }
    historial_queryset = historial_queryset.order_by(ordering_map.get(orden, '-fecha'))

    # Paginación
    page_number = request.GET.get('page')
    paginator = Paginator(historial_queryset, 10)
    historial = paginator.get_page(page_number)

    metodos_pago = (HistorialComanda.objects
                    .exclude(metodo_pago__isnull=True)
                    .exclude(metodo_pago='')
                    .values_list('metodo_pago', flat=True)
                    .distinct()
                    .order_by('metodo_pago'))

    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    extra_query = query_params.urlencode()
    extra_query = f"&{extra_query}" if extra_query else ''

    return render(request, 'core/reportes.html', {
        'historial': historial,
        'total_ventas': total_ventas,
        'filtro': filtro,
        'busqueda': busqueda,
        'orden_actual': orden,
        'metodo_actual': metodo_actual,
        'metodos_pago': metodos_pago,
        'extra_query': extra_query,
    })


@require_POST
@login_required(login_url='login') 
def nueva_comanda(request):
    comanda = Comanda.objects.create(
        fecha=timezone.now(),
        estado='abierta',
        total=0,
        empleado=request.user,
    )
    return redirect('venta')


@csrf_exempt
@require_POST
def verificar_superusuario(request):
    """Verifica permisos de superusuario/dueno.

    - Si hay sesion activa de dueno/superusuario, permite pasar sin volver a pedir
      contrasena.
    - Si no, requiere username + password via authenticate.
    """
    try:
        datos = json.loads(request.body)

        username = (datos.get('username') or '').strip()
        password = (datos.get('password') or '').strip()

        # Atajo: usuario autenticado en sesion y con rol de dueno/superusuario
        if request.user.is_authenticated and _user_is_owner(request.user):
            if (not username and not password) or username == request.user.username:
                return JsonResponse({
                    'es_superusuario': True,
                    'mensaje': 'Sesion de administrador valida'
                })

        # Flujo normal: credenciales explicitas
        if not username or not password:
            return JsonResponse({
                'es_superusuario': False,
                'error': 'Usuario y contrasena son requeridos'
            }, status=400)

        user = authenticate(username=username, password=password)

        if user and _user_is_owner(user):
            return JsonResponse({
                'es_superusuario': True,
                'mensaje': 'Credenciales validas'
            })

        return JsonResponse({
            'es_superusuario': False,
            'error': 'Credenciales incorrectas o no tiene privilegios de administrador'
        }, status=403)

    except Exception as e:
        return JsonResponse({
            'es_superusuario': False,
            'error': str(e)
        }, status=500)


@login_required(login_url='login')
def siguiente_numero_cliente(request):
    return JsonResponse({'siguiente': _next_numero_cliente_servir()})



# Comandas
@require_POST
@login_required(login_url='login')
def agregar_a_comanda(request, comanda_id):
    comanda = get_object_or_404(Comanda, id=comanda_id)
    producto_id = request.POST.get('producto_id')
    cantidad = int(request.POST.get('cantidad', 1))

    producto = get_object_or_404(Producto, id=producto_id)
    subtotal = cantidad * producto.precio

    detalle, creado = DetalleComanda.objects.get_or_create(comanda=comanda, producto=producto)
    if creado:
        detalle.cantidad = cantidad
        detalle.subtotal = subtotal
    else:
        detalle.cantidad += cantidad
        detalle.subtotal += subtotal
    detalle.save()

    comanda.total += subtotal
    comanda.fecha = timezone.now()
    comanda.save()

    return redirect('venta')


@login_required(login_url='login')
def api_usuario_detalle(request, user_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        usuario = User.objects.get(id=user_id)
        data = {
            'id': usuario.id,
            'username': usuario.username,
            'email': usuario.email,
            'date_joined': usuario.date_joined.strftime('%d/%m/%Y %H:%M'),
            'last_login': usuario.last_login.strftime('%d/%m/%Y %H:%M') if usuario.last_login else None,
            'is_active': usuario.is_active
        }
        return JsonResponse(data)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)


# Productos
@login_required(login_url='login')
def productos(request):
    if not _user_is_owner(request.user):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('inicio')

    productos_qs = Producto.objects.all()
    form = ProductoForm()
    categorias = Categoria.objects.all().order_by('nombre')

    busqueda = (request.GET.get('q') or '').strip()
    if busqueda:
        productos_qs = productos_qs.filter(nombre__icontains=busqueda)

    categoria_actual = None
    categoria_param = request.GET.get('categoria')
    if categoria_param:
        try:
            categoria_id = int(categoria_param)
        except (TypeError, ValueError):
            categoria_id = None

        if categoria_id and categorias.filter(id=categoria_id).exists():
            categoria_actual = categoria_id
            productos_qs = productos_qs.filter(categoria_id=categoria_actual)

    orden = request.GET.get('orden') or 'nombre_asc'
    ordering_map = {
        'precio_asc': 'precio',
        'precio_desc': '-precio',
        'nombre_desc': '-nombre',
        'nombre_asc': 'nombre',
    }
    productos_qs = productos_qs.order_by(ordering_map.get(orden, 'nombre'))

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)  
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" agregado exitosamente.')
            return redirect('productos')
        else:
            messages.error(request, 'Error al agregar el producto. Verifica los datos.')

    paginator = Paginator(productos_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    base_query = query_params.urlencode()
    extra_query = f"&{base_query}" if base_query else ''

    return render(request, 'core/productos.html', {
        'productos': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'busqueda': busqueda,
        'orden_actual': orden,
        'categoria_actual': categoria_actual,
        'base_query': base_query,
        'extra_query': extra_query,
        'form': form,
        'categorias': categorias,
    })


@login_required(login_url='login')
def clientes(request):
    if not _user_is_owner(request.user):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('inicio')

    busqueda = (request.GET.get('q') or '').strip()
    clientes_qs = Cliente.objects.prefetch_related('direcciones').order_by('-id')

    if busqueda:
        clientes_qs = clientes_qs.filter(
            Q(nombre__icontains=busqueda)
            | Q(telefono__icontains=busqueda)
            | Q(direcciones__direccion__icontains=busqueda)
        ).distinct()

    return render(request, 'core/clientes.html', {
        'clientes': clientes_qs,
        'busqueda': busqueda,
    })


@require_POST
@login_required(login_url='login')
def crear_cliente(request):
    if not _user_is_owner(request.user):
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    nombre = (request.POST.get('nombre') or '').strip()
    telefono = (request.POST.get('telefono') or '').strip()
    direcciones_raw = request.POST.getlist('direcciones[]')
    direcciones = [direccion.strip() for direccion in direcciones_raw if direccion and direccion.strip()]

    if not nombre:
        return JsonResponse({'success': False, 'message': 'El nombre del cliente es obligatorio'}, status=400)

    if not telefono:
        return JsonResponse({'success': False, 'message': 'El telefono del cliente es obligatorio'}, status=400)

    if Cliente.objects.filter(telefono=telefono).exists():
        return JsonResponse({'success': False, 'message': 'Ya existe un cliente con ese numero de telefono'}, status=400)

    if not direcciones:
        return JsonResponse({'success': False, 'message': 'Debe ingresar al menos una direccion'}, status=400)

    cliente = Cliente.objects.create(nombre=nombre, telefono=telefono)
    DireccionCliente.objects.bulk_create([
        DireccionCliente(cliente=cliente, direccion=direccion)
        for direccion in direcciones
    ])

    return JsonResponse({'success': True, 'cliente_id': cliente.id})


@require_POST
@login_required(login_url='login')
def editar_cliente(request, cliente_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    cliente = get_object_or_404(Cliente, id=cliente_id)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)

    nombre = (data.get('nombre') or '').strip()
    telefono = (data.get('telefono') or '').strip()
    direcciones_raw = data.get('direcciones') or []
    direcciones = [str(direccion).strip() for direccion in direcciones_raw if str(direccion).strip()]

    if not nombre:
        return JsonResponse({'success': False, 'message': 'El nombre del cliente es obligatorio'}, status=400)

    if not telefono:
        return JsonResponse({'success': False, 'message': 'El telefono del cliente es obligatorio'}, status=400)

    if Cliente.objects.filter(telefono=telefono).exclude(id=cliente.id).exists():
        return JsonResponse({'success': False, 'message': 'Ya existe un cliente con ese numero de telefono'}, status=400)

    if not direcciones:
        return JsonResponse({'success': False, 'message': 'Debe ingresar al menos una direccion'}, status=400)

    cliente.nombre = nombre
    cliente.telefono = telefono
    cliente.save(update_fields=['nombre', 'telefono'])

    DireccionCliente.objects.filter(cliente=cliente).delete()
    DireccionCliente.objects.bulk_create([
        DireccionCliente(cliente=cliente, direccion=direccion)
        for direccion in direcciones
    ])

    return JsonResponse({'success': True})


@require_POST
@login_required(login_url='login')
def eliminar_cliente(request, cliente_id):
    if not _user_is_owner(request.user):
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    cliente.delete()
    return JsonResponse({'success': True})

@csrf_exempt
@login_required
def crear_categoria(request):
    if request.method == 'POST':
        if not _user_is_owner(request.user):
            messages.error(request, "No tienes permisos para crear categorías.")
            return redirect('inicio')
        nombre = request.POST.get('nombre_categoria')
        if nombre:
            Categoria.objects.get_or_create(nombre=nombre)
    return redirect('productos')

@csrf_exempt
@login_required
def eliminar_categoria(request, categoria_id):
    if request.method == 'POST':
        if not _user_is_owner(request.user):
            messages.error(request, "No tienes permisos para eliminar categorías.")
            return redirect('inicio')
        categoria = get_object_or_404(Categoria, id=categoria_id)
        categoria.delete()
    return redirect('productos')


@login_required(login_url='login')
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    producto.delete()
    return redirect('productos')

@require_POST
@login_required(login_url='login')
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    
    producto.nombre = request.POST.get('nombre')
    producto.precio = request.POST.get('precio')
    producto.descripcion = request.POST.get('descripcion')

    categoria_id = request.POST.get('categoria')
    if categoria_id:
        producto.categoria = Categoria.objects.get(id=categoria_id)
    else:
        producto.categoria = None

    borrar_imagen = request.POST.get('borrar_imagen')
    if borrar_imagen == '1':
        # Poner la imagen por defecto (o borrar la imagen actual)
        producto.imagen.delete(save=False)  # Borra archivo actual (opcional)
        producto.imagen = None  # Limpia campo para que use default en la plantilla

    elif 'imagen' in request.FILES:
        producto.imagen = request.FILES['imagen']

    producto.save()
    return redirect('productos')

@csrf_exempt
@login_required(login_url='login')
def guardar_comanda(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)

    try:
        detalles_payload, total = _build_detalle_payload(data.get('productos') or [])
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    estado = _sanitize_estado(data.get('estado'))
    tipo_servicio = _sanitize_tipo_servicio(data.get('tipo_servicio'))
    try:
        cliente_nombre, telefono_cliente, direccion_cliente, numero_cliente = _resolve_cliente_data(data, tipo_servicio)
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)
    nota = data.get('nota_comanda')

    montos = {
        'monto_efectivo': data.get('monto_efectivo'),
        'monto_tarjeta_debito': data.get('monto_tarjeta_debito'),
        'monto_tarjeta_credito': data.get('monto_tarjeta_credito'),
        'monto_transferencia': data.get('monto_transferencia'),
    }

    try:
        metodo_pago, payment_breakdown = calculate_payment_breakdown(
            data.get('metodo_pago'),
            total,
            montos,
        )
    except PaymentError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    comanda = Comanda.objects.create(
        nombre_cliente=cliente_nombre,
        telefono_cliente=telefono_cliente,
        direccion_cliente=direccion_cliente,
        numero_cliente=numero_cliente,
        estado=estado,
        total=total,
        empleado=request.user,
        metodo_pago=metodo_pago,
        tipo_servicio=tipo_servicio,
        nota=nota,
        **payment_breakdown,
    )

    DetalleComanda.objects.bulk_create([
        DetalleComanda(
            comanda=comanda,
            producto=detalle['producto'],
            cantidad=detalle['cantidad'],
            subtotal=detalle['subtotal'],
            nota=detalle['nota'],
        )
        for detalle in detalles_payload
    ])

    return JsonResponse({'status': 'ok', 'comanda_id': comanda.id})


@login_required(login_url='login')
def comandas_json(request):
    comandas = Comanda.objects.all().order_by('-id')
    data = [{
        'id': c.id,
        'cliente': c.nombre_cliente or 'No asignado',
        'telefono_cliente': c.telefono_cliente,
        'direccion_cliente': c.direccion_cliente,
        'numero_cliente': c.numero_cliente,
        'estado': c.estado,
        'tipo_servicio': c.tipo_servicio,
        'es_historial': False,
    } for c in comandas]
    return JsonResponse({'comandas': data})


# editar - eliminar comanda


@login_required
def comanda_detalle(request, comanda_id):
    comanda = get_object_or_404(Comanda, id=comanda_id)
    detalles = DetalleComanda.objects.filter(comanda=comanda)

    data = {
        'id': comanda.id,
        'cliente': comanda.nombre_cliente or '',
        'telefono_cliente': comanda.telefono_cliente or '',
        'direccion_cliente': comanda.direccion_cliente or '',
        'numero_cliente': comanda.numero_cliente,
        'total': comanda.total,
        'metodo_pago': comanda.metodo_pago,
        'tipo_servicio': comanda.tipo_servicio,
        'monto_efectivo': comanda.monto_efectivo,
        'monto_tarjeta_debito': comanda.monto_tarjeta_debito,
        'monto_tarjeta_credito': comanda.monto_tarjeta_credito,
        'monto_transferencia': comanda.monto_transferencia,
        'nota_comanda': comanda.nota or '',
        'detalles': [
            {
                'producto': d.producto.nombre,
                'producto_id': d.producto.id,
                'cantidad': d.cantidad,
                'subtotal': d.subtotal,
                'nota': d.nota or '',
                'imagen': d.producto.imagen.url if d.producto.imagen else ''
            } for d in detalles
        ]

    }
    return JsonResponse(data)

@login_required
def historial_comanda_detalle(request, comanda_id):
    comanda = get_object_or_404(HistorialComanda, id=comanda_id)
    detalles = json.loads(comanda.detalles) if comanda.detalles else []

    data = {
        'id': comanda.id,
        'cliente': comanda.nombre_cliente or '',
        'telefono_cliente': getattr(comanda, 'telefono_cliente', '') or '',
        'direccion_cliente': getattr(comanda, 'direccion_cliente', '') or '',
        'numero_cliente': getattr(comanda, 'numero_cliente', None),
        'total': comanda.total,
        'metodo_pago': comanda.metodo_pago,
        'tipo_servicio': comanda.tipo_servicio,
        'monto_efectivo': comanda.monto_efectivo,
        'monto_tarjeta_debito': comanda.monto_tarjeta_debito,
        'monto_tarjeta_credito': comanda.monto_tarjeta_credito,
        'monto_transferencia': comanda.monto_transferencia,
        'cerrado_por': comanda.cerrado_por.username if comanda.cerrado_por else '',
        'fecha': comanda.fecha.strftime('%d/%m/%Y %H:%M') if comanda.fecha else '',
        'nota_comanda': getattr(comanda, 'nota', ''),  # <-- agrega esto si tienes el campo
        'detalles': detalles,
    }
    return JsonResponse(data)

@require_http_methods(["POST"])
@login_required
def eliminar_comanda(request, comanda_id):
    try:
        data = json.loads(request.body)
        motivo = data.get('motivo')

        if not motivo:
            return JsonResponse({'status': 'error', 'message': 'Debe ingresar un motivo'}, status=400)

        comanda = get_object_or_404(Comanda, id=comanda_id)

        EliminacionComanda.objects.create(
            comanda_id=comanda.id,
            nombre_cliente=comanda.nombre_cliente,
            total=comanda.total,
            motivo=motivo,
            eliminado_por=request.user
        )

        comanda.delete()
        return JsonResponse({'status': 'ok'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
@login_required
def editar_comanda(request, id):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)

    comanda = get_object_or_404(Comanda, pk=id)

    try:
        detalles_payload, total = _build_detalle_payload(data.get('productos') or [])
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    montos = {
        'monto_efectivo': data.get('monto_efectivo'),
        'monto_tarjeta_debito': data.get('monto_tarjeta_debito'),
        'monto_tarjeta_credito': data.get('monto_tarjeta_credito'),
        'monto_transferencia': data.get('monto_transferencia'),
    }

    try:
        metodo_pago, payment_breakdown = calculate_payment_breakdown(
            data.get('metodo_pago'),
            total,
            montos,
        )
    except PaymentError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    tipo_servicio = _sanitize_tipo_servicio(data.get('tipo_servicio'))
    try:
        cliente_nombre, telefono_cliente, direccion_cliente, numero_cliente = _resolve_cliente_data(data, tipo_servicio, comanda=comanda)
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    comanda.nombre_cliente = cliente_nombre
    comanda.telefono_cliente = telefono_cliente
    comanda.direccion_cliente = direccion_cliente
    comanda.numero_cliente = numero_cliente
    comanda.estado = _sanitize_estado(data.get('estado'))
    comanda.metodo_pago = metodo_pago
    comanda.tipo_servicio = tipo_servicio
    comanda.total = total
    comanda.nota = data.get('nota_comanda')
    for field, value in payment_breakdown.items():
        setattr(comanda, field, value)
    comanda.save()

    DetalleComanda.objects.filter(comanda=comanda).delete()
    DetalleComanda.objects.bulk_create([
        DetalleComanda(
            comanda=comanda,
            producto=detalle['producto'],
            cantidad=detalle['cantidad'],
            subtotal=detalle['subtotal'],
            nota=detalle['nota'],
        )
        for detalle in detalles_payload
    ])

    return JsonResponse({'status': 'ok'})

    
@csrf_exempt
@login_required(login_url='login')
def cerrar_caja(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

    if not _user_is_owner(request.user):
        return JsonResponse({'status': 'error', 'message': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)

    admin_username = data.get('admin_username') or request.user.username
    try:
        admin_user = User.objects.get(username=admin_username)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Administrador no encontrado'}, status=404)

    if Comanda.objects.filter(estado='abierta').exists():
        return JsonResponse({
            'status': 'error',
            'message': 'No se puede cerrar la caja: hay comandas abiertas. Debe cerrarlas o eliminarlas primero.',
        }, status=400)

    comandas_cerradas = list(Comanda.objects.filter(estado='cerrada').select_related('empleado'))
    if not comandas_cerradas:
        return JsonResponse({'status': 'error', 'message': 'No hay comandas cerradas para procesar'}, status=400)

    historial_entries = []
    for comanda in comandas_cerradas:
        detalles = DetalleComanda.objects.filter(comanda=comanda)
        lista_productos = [
            {
                'producto': detalle.producto.nombre,
                'cantidad': detalle.cantidad,
                'subtotal': detalle.subtotal,
                'nota': detalle.nota or '',
            }
            for detalle in detalles
        ]

        historial_entries.append(HistorialComanda(
            fecha=comanda.fecha,
            nombre_cliente=comanda.nombre_cliente,
            telefono_cliente=comanda.telefono_cliente,
            direccion_cliente=comanda.direccion_cliente,
            numero_cliente=comanda.numero_cliente,
            empleado=comanda.empleado,
            total=comanda.total,
            metodo_pago=comanda.metodo_pago,
            monto_efectivo=comanda.monto_efectivo,
            monto_tarjeta_debito=comanda.monto_tarjeta_debito,
            monto_tarjeta_credito=comanda.monto_tarjeta_credito,
            monto_transferencia=comanda.monto_transferencia,
            tipo_servicio=comanda.tipo_servicio,
            detalles=json.dumps(lista_productos),
            cerrado_por=admin_user,
        ))

    HistorialComanda.objects.bulk_create(historial_entries)
    Comanda.objects.all().delete()

    return JsonResponse({'status': 'ok'})

@csrf_exempt
def verificar_y_eliminar_comanda(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        username = data.get('admin')
        password = data.get('pass')
        motivo = data.get('motivo')
        comanda_id = data.get('comanda_id')

        # Verificar usuario y contrasena o usar sesion de dueno/superusuario
        session_user = request.user if request.user.is_authenticated else None
        user = None

        username = (username or '').strip()
        password = (password or '').strip()

        # Si la sesion actual es de dueno/superusuario, permitir cuando:
        # - no se envian credenciales, o
        # - se envia su propio username sin contrasena (campo autocompletado visual)
        if session_user and _user_is_owner(session_user) and (
            (not username and not password) or
            (username == session_user.username and not password)
        ):
            user = session_user
        else:
            user = authenticate(username=username, password=password)

        if user is None or (user.username != 'admin' and not _user_is_owner(user)):
            return JsonResponse({'success': False, 'error': 'Credenciales incorrectas o sin permisos.'})

        try:
            comanda = Comanda.objects.get(id=comanda_id)
        except Comanda.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Comanda no encontrada.'})

        # Guardar registro de eliminación
        EliminacionComanda.objects.create(
            comanda_id=comanda.id,
            nombre_cliente=comanda.nombre_cliente,
            total=comanda.total,
            motivo=motivo,
            eliminado_por=user
        )

        comanda.delete()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

@csrf_exempt
def verificar_admin_y_eliminar_comandas_eliminadas(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        session_user = request.user if request.user.is_authenticated else None
        user = None

        username = (username or '').strip()
        password = (password or '').strip()

        if session_user and _user_is_owner(session_user) and (
            (not username and not password) or
            (username == session_user.username and not password)
        ):
            user = session_user
        else:
            user = authenticate(username=username, password=password)

        if user is None or (user.username != 'admin' and not _user_is_owner(user)):
            return JsonResponse({'success': False, 'error': 'Credenciales incorrectas o sin permisos.'})

        EliminacionComanda.objects.all().delete()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

# Lista de meses en español
MESES_ES = [
    (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
    (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
    (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre")
]


def _get_available_years():
    """Devuelve todos los años presentes en el historial (o el año actual si no hay datos)."""
    extremos = HistorialComanda.objects.aggregate(
        min_fecha=Min('fecha'),
        max_fecha=Max('fecha'),
    )
    hoy = timezone.localdate()

    min_year = extremos['min_fecha'].year if extremos['min_fecha'] else hoy.year
    max_year = extremos['max_fecha'].year if extremos['max_fecha'] else hoy.year

    # Asegurar que al menos el año actual esté disponible
    min_year = min(min_year, hoy.year)
    max_year = max(max_year, hoy.year)

    return list(range(min_year, max_year + 1))


def _build_excel_report(reportes, sheet_title, filename):
    """Genera el Excel con el historial solicitado y lo retorna como respuesta HTTP."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    encabezados = [
        'ID', 'Fecha', 'Empleado', 'Cliente', 'Método de pago',
        'Monto efectivo', 'Monto tarjeta debito', 'Monto tarjeta credito',
        'Monto transferencia', 'Tipo de servicio', 'Cerrado por', 'Total'
    ]
    ws.append(encabezados)

    total_general = 0
    for reporte in reportes:
        fila = [
            reporte.id,
            reporte.fecha.strftime('%d/%m/%Y %H:%M') if reporte.fecha else '',
            reporte.empleado.username if reporte.empleado else '',
            reporte.nombre_cliente or '',
            reporte.metodo_pago or '',
            reporte.monto_efectivo or 0,
            reporte.monto_tarjeta_debito or 0,
            reporte.monto_tarjeta_credito or 0,
            reporte.monto_transferencia or 0,
            reporte.tipo_servicio or '',
            reporte.cerrado_por.username if reporte.cerrado_por else '',
            reporte.total or 0,
        ]
        ws.append(fila)
        total_general += reporte.total or 0

    ws.append([])

    relleno = [''] * (len(encabezados) - 2)
    fila_total = [*relleno, 'Total General:', total_general]
    ws.append(fila_total)

    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def exportar_reportes_mes(request):
    hoy = datetime.today()
    exportar_mes = request.GET.get('exportar') == '1'
    exportar_historico = request.GET.get('exportar_historico') == '1'

    try:
        mes = int(request.GET.get('mes', hoy.month))
        año = int(request.GET.get('año', hoy.year))
    except ValueError:
        mes = hoy.month
        año = hoy.year

    if not exportar_mes and not exportar_historico:
        años = _get_available_years()
        context = {
            'meses': MESES_ES,
            'años': años,
            'mes_seleccionado': mes,
            'año_seleccionado': año,
            'mes_nombre': dict(MESES_ES).get(mes, mes),
            'dia_seleccionado': hoy.day,
        }
        return render(request, 'core/exportar_reportes_mes.html', context)

    if exportar_historico:
        reportes = HistorialComanda.objects.all().order_by('fecha')
        timestamp = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
        filename = f"reportes_historico_{timestamp}.xlsx"
        return _build_excel_report(reportes, 'Reportes_Historico', filename)

    reportes = HistorialComanda.objects.filter(fecha__year=año, fecha__month=mes).order_by('fecha')
    nombre_mes = dict(MESES_ES).get(mes, f"{mes:02d}")
    filename = f"reportes_{nombre_mes}_{año}.xlsx"
    return _build_excel_report(reportes, f"Reportes_{año}_{mes:02d}", filename)


def exportar_reportes_anual(request):
    hoy = datetime.today()
    año_actual = datetime.now().year
    
    try:
        año = int(request.GET.get('año', hoy.year))
    except ValueError:
        año = hoy.year

    if request.GET.get('exportar_anual') != '1':
        años = _get_available_years()
        context = {
            'años': años,
            'año_seleccionado': año,
        }
        return render(request, 'core/exportar_reportes_anual.html', context)

    reportes = HistorialComanda.objects.filter(fecha__year=año).order_by('fecha')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reportes_{año}"

    encabezados = ['ID', 'Fecha', 'Empleado', 'Cliente', 'Método de pago', 'Monto efectivo', 'Monto tarjeta debito','Monto tarjeta credito', 'Monto transferencia','Tipo de servicio', 'Cerrado por', 'Total'] 
    ws.append(encabezados)

    total_general = 0
    for reporte in reportes:
        fila = [
            reporte.id,
            reporte.fecha.strftime('%d/%m/%Y %H:%M') if reporte.fecha else '',
            reporte.empleado.username if reporte.empleado else '',
            reporte.nombre_cliente or '',
            reporte.metodo_pago or '',
            reporte.monto_efectivo or 0,
            reporte.monto_tarjeta_debito or 0,
            reporte.monto_tarjeta_credito or 0,
            reporte.monto_transferencia or 0,
            reporte.tipo_servicio or '',
            reporte.cerrado_por.username if reporte.cerrado_por else '',
            reporte.total or 0,
        ]
        ws.append(fila)
        total_general += reporte.total or 0

    ws.append([])  # fila vacía

    fila_total = ['','','','','','','','','Total General:', total_general]
    ws.append(fila_total)

    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font

    filename = f"reportes_anual_{año}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def imprimir_boleta_thermal(comanda):
    hoy = timezone.localdate()
    n_diario = Comanda.objects.filter(fecha__date=hoy, estado='cerrada', id__lte=comanda.id).count()

    detalles = DetalleComanda.objects.filter(comanda=comanda)
    usuario = comanda.empleado.username if comanda.empleado else ''
    # Ajusta la fecha/hora a la zona horaria activa (Chile)
    fecha_chile = timezone.localtime(comanda.fecha)
    fecha_hora = fecha_chile.strftime('%d/%m/%Y %H:%M')

    ticket = ""
    ticket += "        Punto de Venta\n"
    ticket += "-"*32 + "\n"
    ticket += f"N interno: {comanda.id}\n"
    ticket += f"N diario: {n_diario}\n"
    ticket += f"Responsable: {usuario}\n"
    ticket += f"Fecha: {fecha_hora}\n"
    ticket += "-"*32 + "\n"
    # Eliminar encabezado de cantidad, orden y precio
    # ticket += f"{'Cant':<4} {'Orden':<20} {'Precio':>8}\n"
    # ticket += "-"*40 + "\n"
    max_width = 32
    for det in detalles:
        price_str = f"${det.subtotal:,.0f}"
        prod_str = f"{det.cantidad} x {det.producto.nombre}"
        avail_width = max_width - len(price_str) - 1
        if len(prod_str) > avail_width:
            # Buscar el último espacio antes del límite para no cortar palabras
            cut = prod_str[:avail_width].rfind(' ')
            if cut == -1:
                cut = avail_width
            first_line = prod_str[:cut]
            rest = prod_str[cut:].lstrip()
            ticket += f"{first_line:<{avail_width}}{price_str:>{len(price_str)+1}}\n"
            # Resto del nombre en líneas siguientes, sin cortar palabras
            words = rest.split()
            line = ''
            for word in words:
                if len(line) + len(word) + (1 if line else 0) <= max_width:
                    line += (' ' if line else '') + word
                else:
                    ticket += f"{line}\n"
                    line = word
            if line:
                ticket += f"{line}\n"
        else:
            ticket += f"{prod_str:<{avail_width}} {price_str:>{len(price_str)}}\n"
        if det.nota:
            ticket += f"  * Nota: {det.nota}\n"
    ticket += "-"*32 + "\n"
    if getattr(comanda, 'nota', None):
        ticket += f"Nota: {comanda.nota}\n"
        ticket += "-"*32 + "\n"
    ticket += f"{'Total:':<25}"
    ticket += f"${comanda.total:,.0f}\n"
    ticket += "iva incluido\n"
    ticket += "-"*32 + "\n"
    ticket += f"{comanda.tipo_servicio}, {comanda.nombre_cliente}\n"
    ticket += "\n\n"

    try:
        send_raw_to_printer("CAJA", f"Boleta #{comanda.id}", ticket)
    except PrinterError:
        logger.exception("Error al imprimir boleta thermal de la comanda %s", comanda.id)
        raise

def comandas_eliminadas(request):
    if not _user_is_owner(request.user):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('inicio')

    todas = EliminacionComanda.objects.all().order_by('-fecha_eliminacion')
    paginator = Paginator(todas, 16)  # 50 por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/comandas_eliminadas.html', {
        'comandas': page_obj,
        'es_owner_actual': _user_is_owner(request.user),
    })

@login_required(login_url='login')
def comandas_eliminadas_json(request):
    if not _user_is_owner(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    data = []
    for e in EliminacionComanda.objects.select_related('eliminado_por'):
        data.append({
            "comanda_id": e.comanda_id,
            "nombre_cliente": e.nombre_cliente,
            "total": e.total,
            "motivo": e.motivo,
            "fecha_eliminacion": e.fecha_eliminacion.strftime("%d/%m/%Y %H:%M"),
            "eliminado_por_nombre": e.eliminado_por.username if e.eliminado_por else "Desconocido",
        })
    return JsonResponse({"eliminadas": data})


@require_POST
@login_required(login_url='login')
def reabrir_comanda(request, comanda_id):
    comanda = get_object_or_404(Comanda, id=comanda_id, estado='cerrada')
    comanda.estado = 'abierta'
    comanda.save()
    return JsonResponse({'status': 'ok'})



def imprimir_comanda_cocina(comanda):
    hoy = timezone.localdate()
    n_diario = Comanda.objects.filter(fecha__date=hoy, estado='cerrada', id__lte=comanda.id).count()
    detalles = DetalleComanda.objects.filter(comanda=comanda)
    usuario = comanda.empleado.username if comanda.empleado else ''
    fecha_chile = timezone.localtime(comanda.fecha)
    fecha_hora = fecha_chile.strftime('%d/%m/%Y %H:%M')

    ticket = ""
    ticket += f"N interno: {comanda.id}\n"
    ticket += f"N diario: {n_diario}\n"
    ticket += f"Fecha: {fecha_hora}\n"
    ticket += f"Responsable: {usuario}\n"
    ticket += "-"*32 + "\n"
    max_width = 32
    for det in detalles:
        prod_str = f"{det.cantidad} x {det.producto.nombre}"
        # Print product name, wrap if too long
        rest = prod_str
        while rest:
            ticket += f"{rest[:max_width]}\n"
            rest = rest[max_width:]
        if det.nota:
            ticket += f"  * Nota: {det.nota}\n"
    ticket += "-"*32 + "\n"
    if getattr(comanda, 'nota', None):
        ticket += f"Nota: {comanda.nota}\n"
        ticket += "-"*32 + "\n"
    ticket += f"{comanda.tipo_servicio}, {comanda.nombre_cliente}\n"
    ticket += "\n\n"

    try:
        send_raw_to_printer("COCINA", f"Comanda #{comanda.id}", ticket)
    except PrinterError:
        logger.exception("Error al imprimir la comanda de cocina %s", comanda.id)
        raise        


@require_POST
@login_required(login_url='login')
def imprimir_boleta_ventas_dia(request):
    hoy = timezone.localdate()

    ventas = HistorialComanda.objects.filter(fecha__date=hoy)
    total_venta = 0
    total_pedidos = ventas.count()
    metodos = {}

    for v in ventas:
        metodo = (v.metodo_pago or "Otros").strip().lower()
        if metodo == "mixto":
            if hasattr(v, "monto_efectivo") and v.monto_efectivo:
                metodos["Efectivo"] = metodos.get("Efectivo", 0) + v.monto_efectivo
            if hasattr(v, "monto_tarjeta_debito") and v.monto_tarjeta_debito:
                metodos["Tarjeta Debito"] = metodos.get("Tarjeta Debito", 0) + v.monto_tarjeta_debito
            if hasattr(v, "monto_tarjeta_credito") and v.monto_tarjeta_credito:
                metodos["Tarjeta Credito"] = metodos.get("Tarjeta Credito", 0) + v.monto_tarjeta_credito
            if hasattr(v, "monto_transferencia") and v.monto_transferencia:
                metodos["Transferencia"] = metodos.get("Transferencia", 0) + v.monto_transferencia
        elif metodo == "efectivo":
            metodos["Efectivo"] = metodos.get("Efectivo", 0) + (getattr(v, "monto_efectivo", 0) or v.total or 0)
        elif metodo == "tarjeta_debito":
            metodos["Tarjeta Debito"] = metodos.get("Tarjeta Debito", 0) + (getattr(v, "monto_tarjeta_debito", 0) or v.total or 0)
        elif metodo == "tarjeta_credito":
            metodos["Tarjeta Credito"] = metodos.get("Tarjeta Credito", 0) + (getattr(v, "monto_tarjeta_credito", 0) or v.total or 0)
        elif metodo == "transferencia":
            metodos["Transferencia"] = metodos.get("Transferencia", 0) + (getattr(v, "monto_transferencia", 0) or v.total or 0)
        else:
            nombre = v.metodo_pago.strip().title() if v.metodo_pago else "Otros"
            metodos[nombre] = metodos.get(nombre, 0) + v.total
        total_venta += v.total

    usuario = request.user.username if request.user.is_authenticated else 'Sistema'
    fecha_hora = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')

    ticket = ""
    ticket += f"Resumen de venta: {hoy.strftime('%d/%m/%Y')}\n"
    ticket += "-"*32 + "\n"
    ticket += "Resumen de Boletas, Transbank y otros:\n"
    for metodo, monto in metodos.items():
        ticket += f"{metodo:<22}${monto:,.0f}\n"
    ticket += "-"*32 + "\n"
    ticket += f"Total venta: ${total_venta:,.0f}\n"
    ticket += f"{'Total de pedidos del dia:':<30}{total_pedidos}\n"
    ticket += "-"*32 + "\n"
    # Centrar la fecha al final
    ticket += f"{fecha_hora:^{32}}\n"
    ticket += "\n\n"

    try:
        send_raw_to_printer("CAJA", "Ventas del Día", ticket)
        return HttpResponse("Boleta de ventas del día enviada a la impresora.")
    except PrinterError:
        logger.exception("Error al imprimir el resumen de ventas del día")
        return HttpResponse("No se pudo enviar el resumen a la impresora.", status=500)

@csrf_exempt
@login_required(login_url='login')
def imprimir_boleta_comanda(request, comanda_id):
    if request.method == 'POST':
        try:
            comanda = Comanda.objects.get(id=comanda_id)
            imprimir_boleta_thermal(comanda)
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})

@csrf_exempt
@login_required(login_url='login')
def imprimir_boleta_cocina_view(request, comanda_id):
    if request.method == 'POST':
        try:
            comanda = Comanda.objects.get(id=comanda_id)
            imprimir_comanda_cocina(comanda)
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})